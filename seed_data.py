"""
One-shot script: parse OT酒单.xlsx → data/cocktails.json
"""
import json, os, re
from openpyxl import load_workbook

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'cocktails.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'OT酒单.xlsx')
WEB_IMAGES = os.path.join(os.path.dirname(__file__), 'static', 'web_images')

# ---- spirit category mapping ----
SPIRIT_MAP = {
    '金酒': 'gin', '威士忌': 'whisky', '伏特加': 'vodka',
    '白兰地': 'brandy', '朗姆': 'rum', '龙舌兰': 'tequila', '其他': 'other',
}

# ---- glassType by name keyword ----
def infer_glass_type(glass_name):
    name = glass_name.lower()
    if '马天尼' in name: return 'martini'
    if 'colin' in name: return 'highball'
    if '香槟' in name: return 'coupe'
    if '威士忌' in name: return 'rocks'
    if '勃艮第' in name: return 'brandySnifter'
    if '郁金香' in name: return 'coupe'
    if '玛格丽特' in name: return 'margarita'
    if '飓风' in name: return 'hurricane'
    return 'highball'

# ---- color by spirit ----
SPIRIT_COLORS = {
    'gin': '#c8e6c9', 'whisky': '#ffcc80', 'vodka': '#e0e0e0',
    'rum': '#ffe0b2', 'brandy': '#d4a574', 'tequila': '#c8e6c9', 'other': '#f8bbd0',
}

# ---- guess spirit from ingredient text ----
def guess_spirit(desc):
    d = desc or ''
    if '金酒' in d or 'Gin' in d: return 'gin'
    if '威士忌' in d or '波本' in d or 'Whisky' in d or 'Whiskey' in d: return 'whisky'
    if '伏特加' in d or 'Vodka' in d: return 'vodka'
    if '白兰地' in d or '干邑' in d or 'Brandy' in d or 'Cognac' in d: return 'brandy'
    if '朗姆' in d or 'Rum' in d: return 'rum'
    if '龙舌兰' in d or '梅斯卡尔' in d or 'Tequila' in d or 'Agave' in d: return 'tequila'
    return 'other'

# ---- check if image exists ----
def find_image(name):
    """Check if a JPG/PNG image exists for this drink name."""
    for ext in ('.jpg', '.jpeg', '.png', '.webp'):
        fname = name + ext
        path = os.path.join(WEB_IMAGES, fname)
        if os.path.exists(path):
            return 'web_images/' + fname
    return None

def parse_name_price(text):
    """Parse '藤井树\nLove letter\n¥96' → (name, en, price)"""
    if not text:
        return '', '', ''
    lines = text.strip().split('\n')
    name = lines[0].strip() if len(lines) > 0 else ''
    en = ''
    price = ''
    for line in lines[1:]:
        s = line.strip()
        if s.startswith('¥') or s.startswith('¥'):
            price = s
        elif s and not s[0].isdigit():
            en = s
    # also check for price pattern in name line
    m = re.search(r'¥\d+', text)
    if m and not price:
        price = m.group()
    return name, en, price

# ===========================
# PARSE SHEET 1: 2026春 (specials)
# ===========================
def parse_specials(ws):
    drinks = []
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        col1 = str(row[0]).strip() if row[0] is not None else ''
        # detect new drink: first column is a number (1-12)
        if col1 and col1[0].isdigit():
            if current:
                drinks.append(current)
            name, en, price = parse_name_price(str(row[1]) if row[1] else '')
            abv = str(row[2]).strip() if row[2] is not None else ''
            desc = str(row[4]).strip() if row[4] is not None else ''
            flavor = str(row[5]).strip() if row[5] is not None else ''
            spirit = guess_spirit(str(row[3]) if row[3] else '')
            current = {
                'id': '', 'name': name, 'en': en, 'price': price or '¥96',
                'abv': abv + '%' if abv and not abv.endswith('%') else (abv or '??%'),
                'cat': '特调', 'desc': desc, 'flavor': flavor,
                'glass': '', 'method': '',
                'spirit': spirit, 'color': SPIRIT_COLORS.get(spirit, '#f8bbd0'),
                'glassType': 'highball', 'img': '', 'hasImg': False,
            }
    if current:
        drinks.append(current)

    # assign IDs and glass/method from parsed data
    id_map = {
        '藤井树': 'fjs', '澪': 'mio', '记念': 'jinian', '休假2.0': 'xj2',
        '阿基拉': 'akira', 'AKIRA': 'akira', '毒藤': 'dt', 'R-16玛丽': 'r16',
        '哥伦布斯': 'glbs', '潮红': 'ch', '地平线': 'dpx', '兰纳': 'ln', '芫荽': 'ys',
    }
    # Fix AKIRA name
    for d in drinks:
        if d['name'] == 'AKIRA':
            d['name'] = '阿基拉'
            d['en'] = 'AKIRA'
    # glass and method from original HTML
    glass_method_map = {
        '藤井树': ('矮脚香槟杯', '奶洗', 'coupe'),
        '澪': ('colin杯', 'Shake', 'highball'),
        '记念': ('colin杯', 'shake', 'highball'),
        '休假2.0': ('colin杯', 'shake', 'highball'),
        '阿基拉': ('郁金香杯', 'Shake', 'coupe'),
        '毒藤': ('毒藤杯', 'Shake', 'highball'),
        'R-16玛丽': ('条纹威士忌杯', 'Rolling', 'rocks'),
        '哥伦布斯': ('高脚马天尼杯', 'Shake', 'martini'),
        '潮红': ('矮脚香槟杯', 'Shake', 'coupe'),
        '地平线': ('小威士忌杯', 'Shake', 'rocks'),
        '兰纳': ('大勃艮第杯', 'Shake', 'brandySnifter'),
        '芫荽': ('akira的杯子', 'Shake', 'highball'),
    }

    for d in drinks:
        d['id'] = id_map.get(d['name'], 's' + str(drinks.index(d)))
        if d['name'] in glass_method_map:
            g, m, gt = glass_method_map[d['name']]
            d['glass'] = g
            d['method'] = m
            d['glassType'] = gt
        img = find_image(d['name'])
        if img:
            d['img'] = img
            d['hasImg'] = True

    return drinks

# ===========================
# PARSE SHEET 2: 经典 (classics)
# ===========================
def parse_classics(ws):
    drinks = []
    current_spirit = 'other'
    for row in ws.iter_rows(min_row=2, values_only=True):
        col1 = str(row[0]).strip() if row[0] is not None else ''
        col2 = str(row[1]).strip() if row[1] is not None else ''
        # spirit category header row
        if col1 and not col2:
            current_spirit = SPIRIT_MAP.get(col1.split('\n')[0], 'other')
            continue
        if not col1 and not col2:
            continue
        name, en, _ = parse_name_price(col2)
        abv_raw = str(row[2]).strip() if row[2] is not None else ''
        # Clean "12.0%abv" → "12.0%"
        abv = re.sub(r'%abv.*$', '%', abv_raw)
        if abv and not abv.endswith('%'):
            abv = abv + '%'
        desc_raw = str(row[3]).strip() if row[3] is not None else ''
        desc = desc_raw.split('\n')[0]  # Chinese only
        price_raw = str(row[4]).strip() if row[4] is not None else ''
        price = '¥' + price_raw if price_raw and not price_raw.startswith('¥') else (price_raw or '¥96')

        # glassType defaults
        glass_defaults = {
            '金汤力': 'highball', '飞行': 'martini', '金菲士': 'highball', '琴蕾': 'martini',
            '三叶草俱乐部': 'coupe', '尼格罗尼': 'rocks', '马天尼': 'martini', '遗言': 'coupe',
            '纸飞机': 'coupe', '威士忌酸': 'rocks', '古典': 'rocks', '曼哈顿': 'coupe',
            '教父': 'rocks', '老广场': 'rocks',
            '莫斯科骡子': 'highball', '性感沙滩': 'highball', '神风敢死队': 'martini',
            '黑/白俄罗斯': 'rocks', '血腥玛丽': 'highball', '大都会': 'martini',
            '马颈': 'highball', '边车': 'coupe', '床笫之间': 'coupe', 'B&B': 'brandySnifter',
            '萨泽拉克': 'rocks',
            '大吉利': 'coupe', '莫吉托': 'highball', '椰林飘香': 'hurricane',
            '龙舌兰日出': 'highball', '大魔鬼': 'highball', '玛格丽特': 'margarita',
            '长岛冰茶': 'highball', '青蚱蜢': 'coupe', '杏仁酸': 'rocks',
            '皮斯科酸': 'coupe', '查理卓别林': 'coupe',
        }

        d = {
            'id': '', 'name': name, 'en': en, 'price': price,
            'abv': abv, 'cat': '经典', 'desc': desc, 'flavor': '',
            'glass': '', 'method': '',
            'spirit': current_spirit, 'color': SPIRIT_COLORS.get(current_spirit, '#f8bbd0'),
            'glassType': glass_defaults.get(name, 'highball'),
            'img': '', 'hasImg': False,
        }
        img = find_image(name)
        if img:
            d['img'] = img
            d['hasImg'] = True
        drinks.append(d)

    # assign IDs
    id_map = {
        '金汤力': 'jtl', '飞行': 'fx', '金菲士': 'jfs', '琴蕾': 'ql',
        '三叶草俱乐部': 'sycklb', '尼格罗尼': 'ngln', '马天尼': 'mtn', '遗言': 'yy',
        '纸飞机': 'zfj', '威士忌酸': 'wjs', '古典': 'gd', '曼哈顿': 'mhd',
        '教父': 'jf', '老广场': 'lgc',
        '莫斯科骡子': 'msklz', '性感沙滩': 'xgst', '神风敢死队': 'sfgsd',
        '黑/白俄罗斯': 'bwr', '血腥玛丽': 'xxml', '大都会': 'ddh',
        '马颈': 'mj', '边车': 'bc', '床笫之间': 'czzj', 'B&B': 'bb', '萨泽拉克': 'szlk',
        '大吉利': 'djl', '莫吉托': 'mjt', '椰林飘香': 'ylpx',
        '龙舌兰日出': 'lslrc', '大魔鬼': 'dmg', '玛格丽特': 'mglt',
        '长岛冰茶': 'cdbc', '青蚱蜢': 'qzm', '杏仁酸': 'xrs',
        '皮斯科酸': 'psks', '查理卓别林': 'clzbl',
    }
    for d in drinks:
        d['id'] = id_map.get(d['name'], 'c' + str(drinks.index(d)))
    return drinks

# ===========================
# PARSE SHEET 3: 纯饮 (straight pours)
# ===========================
def parse_pure(ws):
    drinks = []
    current_cat = '纯饮'
    current_type = ''
    for row in ws.iter_rows(min_row=2, values_only=True):
        col1 = str(row[0]).strip() if row[0] is not None else ''
        col2 = str(row[1]).strip() if row[1] is not None else ''
        # category line
        if col1 and not col2:
            current_type = col1
            continue
        if not col2:
            continue
        # skip header hints
        if '杯饮' in col2 or '苏威' in col2 or '日威' in col2 or '金酒' in col1 or '其他' in col1:
            continue

        name, en, _ = parse_name_price(col2)
        # Skip non-drink category headers
        if name in ('金酒', '其他瓶装', '巴黎之花', '818金龙舌兰', '轩尼诗VSOP'):
            continue
        abv_raw = str(row[2]).strip() if row[2] is not None else ''
        # Clean "12.0%abv" → "12.0%"
        abv = re.sub(r'%abv.*$', '%', abv_raw)
        if abv and not abv.endswith('%'):
            abv = abv + '%'
        price_raw = str(row[3]).strip() if row[3] is not None else ''
        price = '¥' + price_raw if price_raw and price_raw != '/' and not price_raw.startswith('¥') else ('' if price_raw == '/' else price_raw)
        note = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''

        drink_type = col1 if col1 else current_type
        d = {
            'id': 'p' + str(len(drinks) + 1),
            'name': name, 'en': en, 'cat': '纯饮',
            'price': price or '¥86', 'abv': abv or '40%',
            'desc': drink_type, 'flavor': '', 'glass': '', 'method': '',
            'spirit': 'other', 'color': '', 'glassType': '',
            'img': '', 'hasImg': False, 'note': note,
        }
        img = find_image(name)
        if img:
            d['img'] = img
            d['hasImg'] = True
        drinks.append(d)
    return drinks

# ===========================
# MAIN
# ===========================
if __name__ == '__main__':
    wb = load_workbook(EXCEL_FILE, data_only=True)

    specials = parse_specials(wb['2026春'])
    classics = parse_classics(wb['经典'])
    pure = parse_pure(wb['纯饮'])

    all_drinks = specials + classics + pure

    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_drinks, f, ensure_ascii=False, indent=2)

    print(f'Total: {len(all_drinks)} drinks')
    print(f'  Specials: {len(specials)}')
    print(f'  Classics: {len(classics)}')
    print(f'  Pure:     {len(pure)}')
    for d in all_drinks:
        img_mark = ' [IMG]' if d['hasImg'] else ''
        print(f'  [{d["cat"]}] {d["name"]} ({d["en"]}) - {d["price"]} {d["abv"]}{img_mark}')
